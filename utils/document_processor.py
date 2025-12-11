import base64
import os
from pdf2image import convert_from_path
from anthropic import Anthropic
from dotenv import load_dotenv
import io
from PIL import Image

load_dotenv()

class DocumentProcessor:
    """Process uploaded documents and extract knowledge using Claude."""

    def __init__(self):
        """Initialize Anthropic client."""
        api_key = os.getenv('ANTHROPIC_API_KEY')
        if not api_key:
            raise ValueError("ANTHROPIC_API_KEY not found in environment variables")
        self.client = Anthropic(api_key=api_key)

    def process_pdf(self, file_path):
        """
        Convert PDF to images and send to Claude for analysis.

        Args:
            file_path: Path to PDF file

        Returns:
            dict: Extracted knowledge from document
        """
        # Convert PDF pages to images
        images = convert_from_path(file_path, dpi=150)

        # Prepare content for Claude with all pages
        content = []
        for i, image in enumerate(images):
            # Convert PIL Image to base64
            buffered = io.BytesIO()
            image.save(buffered, format="PNG")
            img_base64 = base64.b64encode(buffered.getvalue()).decode('utf-8')

            # Add image to content
            content.append({
                "type": "image",
                "source": {
                    "type": "base64",
                    "media_type": "image/png",
                    "data": img_base64
                }
            })

        # Add text prompt after all images
        content.append({
            "type": "text",
            "text": """Analyze this study material and extract all factual information, key concepts, definitions, relationships, and testable knowledge.

Organize your findings by topic/category. Include:
- Key facts and definitions
- Important concepts and their relationships
- Processes and how they work
- Dates, names, and specific details
- Any testable information

Be thorough and precise. This will be used to generate quiz questions."""
        })

        # Call Claude API
        message = self.client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=4096,
            messages=[
                {
                    "role": "user",
                    "content": content
                }
            ]
        )

        # Extract text response
        knowledge = message.content[0].text

        return {
            "type": "pdf",
            "pages": len(images),
            "knowledge": knowledge
        }

    def process_spreadsheet(self, file_path):
        """
        Convert spreadsheet to markdown and send to Claude for analysis.

        Args:
            file_path: Path to Excel file

        Returns:
            dict: Extracted knowledge from document
        """
        from openpyxl import load_workbook

        # Load workbook
        workbook = load_workbook(file_path, data_only=True)

        # Convert all sheets to markdown
        markdown_content = []

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            markdown_content.append(f"## Sheet: {sheet_name}\n")

            # Convert sheet to markdown table
            rows = []
            for row in sheet.iter_rows(values_only=True):
                # Skip empty rows
                if all(cell is None or str(cell).strip() == '' for cell in row):
                    continue
                # Convert row to strings
                row_strings = [str(cell) if cell is not None else '' for cell in row]
                rows.append('| ' + ' | '.join(row_strings) + ' |')

            if rows:
                # Add header separator after first row
                if len(rows) > 1:
                    num_cols = rows[0].count('|') - 1
                    separator = '| ' + ' | '.join(['---'] * num_cols) + ' |'
                    rows.insert(1, separator)

                markdown_content.extend(rows)

            markdown_content.append('\n')

        markdown_text = '\n'.join(markdown_content)

        # Call Claude API
        message = self.client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=4096,
            messages=[
                {
                    "role": "user",
                    "content": f"""Analyze this study material (from a spreadsheet) and extract all factual information, key concepts, definitions, relationships, and testable knowledge.

{markdown_text}

Organize your findings by topic/category. Include:
- Key facts and definitions
- Important concepts and their relationships
- Processes and how they work
- Dates, names, and specific details
- Any testable information

Be thorough and precise. This will be used to generate quiz questions."""
                }
            ]
        )

        # Extract text response
        knowledge = message.content[0].text

        return {
            "type": "spreadsheet",
            "sheets": len(workbook.sheetnames),
            "knowledge": knowledge
        }

    def process_document(self, file_path, file_extension):
        """
        Process document based on file type.

        Args:
            file_path: Path to uploaded file
            file_extension: File extension (pdf, xlsx, xls)

        Returns:
            dict: Extracted knowledge from document
        """
        if file_extension == 'pdf':
            return self.process_pdf(file_path)
        elif file_extension in ['xlsx', 'xls']:
            return self.process_spreadsheet(file_path)
        else:
            raise ValueError(f"Unsupported file type: {file_extension}")
